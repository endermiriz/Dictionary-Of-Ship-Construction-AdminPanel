import time
from datetime import datetime
from PyQt5 import QtWidgets, QtGui
import ctypes
from shareplum import Site
from shareplum import Office365
from shareplum.site import Version
from giris import Ui_Dialog
from PyQt5.QtWidgets import *
from PyQt5.QtCore    import *
from PyQt5.QtGui     import *
from panel import Ui_MainWindow
import pandas as pd
import re
import openpyxl
import itertools
import os

myappid = 'dictionary.of.shipbuilding.adminpanel.0.0.1.version.created.by.Ender.MIRIZ'
ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
excelemail = []
excelpass = []
class Login(QDialog):
    def __init__(self):
        super().__init__()
        self.ui = Ui_Dialog()
        self.ui.setupUi(self)
        self.textName = self.ui.lineEdit
        self.textPass = self.ui.lineEdit_2
        self.buttonLogin = self.ui.pushButton
        self.buttonLogin.clicked.connect(self.handleLogin)
        self.ui.pushButton_2.clicked.connect(self.clean)
        self.setWindowTitle("Naval Dictionary")
        self.setWindowIcon(QIcon("icon\logo.png"))
        self.textPass.setEchoMode(QLineEdit.Password)
        label = self.ui.label_3
        pixmap = QPixmap('icon\logo.png')
        label.setPixmap(pixmap)



    def clean(self):
        self.textName.clear()
        self.textPass.clear()
    def handleLogin(self):
        try:
            authcookie = Office365('https://pirireisedutr.sharepoint.com/', username=self.textName.text(),
                                   password=self.textPass.text()).GetCookies()

            i = 1
        except:
            i = 0
        if i == 1:
            excelemail.append(self.textName.text())
            excelpass.append(self.textPass.text())
            site = Site('https://pirireisedutr.sharepoint.com/sites/GeminaatSzlkProjesiGrubu', version=Version.v2016,
                        authcookie=authcookie)
            folder = site.Folder('Shared Documents/Beta Testing')
            file = folder.get_file('dict_data.xlsx')
            with open("dict_data.xlsx", "wb") as fh:
                fh.write(file)
            with open("dict_data.xlsx", "rb") as file_obj:
                file_as_string = file_obj.read()
            self.accept()
        else:
            QtWidgets.QMessageBox.warning(
                self, 'Hata', 'Giriş bilgileriniz hatalı!\n\nLütfen email ve şifreyi kontrol edip tekrar deneyin.')



class Window(QMainWindow):

    def __init__(self):
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.setWindowTitle("Naval Dictionary -Admin Panel-")
        self.setWindowIcon(QIcon("icon\logo.png"))
        self.setlist()
        self.updatekelimelist()
        self.ui.lineEdit.textChanged.connect(self.findkelime)
        self.ui.pushButton.clicked.connect(self.findexcelkelime)
        self.ui.pushButton_2.clicked.connect(self.clearpanel)
        self.ui.pushButton_6.clicked.connect(self.listedenkaldir)
        self.ui.label_3.setText(*excelemail)
        self.ui.pushButton_3.clicked.connect(self.yukle)
        self.ui.pushButton_5.clicked.connect(self.duzenle)
    def setlist(self):
        df = pd.read_excel('dict_data.xlsx')
        df.Kelimeler = df.Kelimeler.str.capitalize()
        df.Anlamlar = df.Anlamlar.str.capitalize()
        df = df.sort_values(by=['Kelimeler', 'Anlamlar'])
        df = df.drop_duplicates(subset=['Kelimeler'], keep='last')
        # df['Anlamlar'] = df["Kelimeler"] + "\n\n" + df['Anlamlar']
        df.to_excel('dict_data_setting.xlsx',sheet_name="1", index=False)

    Kelimelers = []
    def updatekelimelist(self):
        data = pd.read_excel('dict_data_setting.xlsx')
        kelimeler = data["Kelimeler"].tolist()
        Window.Kelimelers.clear()
        for kelime in kelimeler:
            parantezdisi = kelime.split(' [')[0]
            item = QListWidgetItem(kelime)
            self.ui.listWidget.addItem(item)
            Window.Kelimelers.append(parantezdisi)
    def duzenle(self):

        btn = self.ui.pushButton_5
        texting = btn.text()
        itemd = Window.items
        if self.ui.listWidget_3.selectedItems():
            if texting == "Düzenlemeyi Kaydet":
                listItems = self.ui.listWidget_3.selectedItems()
                qm = QtWidgets.QMessageBox
                for item in listItems:
                    ret = qm.question(self, 'Uyarı',
                                      "'" + item.text() + "'" + " kelimesinin düzenlenmiş halini kaydetmek istediğinizden emin misiniz?",
                                      qm.Yes | qm.No)
                    if ret == qm.Yes:
                        item = self.ui.lineEdit.text()
                        item = item.capitalize()
                        tritem = self.ui.lineEdit_2.text()
                        tritem = tritem.capitalize()
                        tranlam = self.ui.plainTextEdit.toPlainText()
                        inganlam = self.ui.plainTextEdit_2.toPlainText()
                        if len(tritem and item and tranlam and inganlam) == 0 or tritem.isspace() == True:
                            QtWidgets.QMessageBox.warning(
                                self, 'Uyarı', 'Listeye ekleme yapmak için gerekli yerleri doldurunuz!')

                        else:
                            if item in itemd:
                                QtWidgets.QMessageBox.warning(
                                    self, 'Uyarı',
                                    'İstediğiniz kelimeyi zaten listeye eklemişsiniz!\n\nListede var olan bir kelimeyi tekrar ekleyemezsiniz.')

                            else:
                                items = Window.items
                                Kelimeler = Window.Kelimelers
                                item = self.ui.lineEdit.text()
                                item = item.capitalize()
                                if item in Kelimeler:
                                    QtWidgets.QMessageBox.warning(
                                        self, 'Uyarı',
                                        'İstediğiniz kelime zaten Excel Dosyasına kayıtlı!\n\nExcel Dosyasında var olan bir kelimeyi tekrar ekleyemezsiniz.')
                                else:
                                    self.listedenkaldirduzenleme()
                                    btn.setText("Düzenle")
                                    self.ui.pushButton_6.setEnabled(True)
                                    self.ui.pushButton_3.setEnabled(True)
                                    self.ui.pushButton.setEnabled(True)
                                    self.ui.pushButton_2.setEnabled(True)
                                    itemler = self.ui.listWidget_3
                                    itemler.setEnabled(True)
                                    self.addkelimelist()
            else:
                if self.ui.listWidget_3.selectedItems():
                    birlesikkelime = Window.birlesikkelime
                    if self.ui.listWidget_3.selectedItems():
                        listItems = self.ui.listWidget_3.selectedItems()
                        qm = QtWidgets.QMessageBox
                        for item in listItems:
                            ret = qm.question(self, 'Uyarı',
                                              "'" + item.text() + "'" + " kelimesini düzenlemek istediğinizden istediğinizden emin misiniz?",
                                              qm.Yes | qm.No)
                            if ret == qm.Yes:
                                secilen = self.ui.listWidget_3.selectedItems()
                                for sec in secilen:
                                    sec = sec.text()
                                    item = str(sec)
                                    parantezdisi = item.split(' [')[0]
                                    Window.items.remove(parantezdisi)


                                items = self.ui.listWidget_3
                                items.setEnabled(False)
                                self.ui.pushButton_6.setEnabled(False)
                                self.ui.pushButton_3.setEnabled(False)
                                self.ui.pushButton.setEnabled(False)
                                self.ui.pushButton_2.setEnabled(False)
                                if btn is not None:
                                    text = btn.text()
                                    btn.setText("Düzenlemeyi Kaydet" if text == "Düzenle" else "Düzenle")
                                    self.checkduzenleme()
                                if not listItems: return
                                for item in listItems:
                                    pass

    items = []
    tritems = []
    tringanlam = []

    def addkelimelist(self):
        items = Window.items
        item = self.ui.lineEdit.text()
        item = item.capitalize()

        tritems = Window.tritems
        tritem = self.ui.lineEdit_2.text()
        tritem = tritem.capitalize()
        tranlam = self.ui.plainTextEdit.toPlainText()
        inganlam = self.ui.plainTextEdit_2.toPlainText()
        if len(tritem+item+tranlam+inganlam) == 0 or tritem.isspace() == True:
            QtWidgets.QMessageBox.warning(
                self, 'Uyarı', 'Listeye ekleme yapmak için gerekli yerleri doldurunuz!')
        else:
                    if item in items:
                        QtWidgets.QMessageBox.warning(
                            self, 'Uyarı', 'İstediğiniz kelimeyi zaten listeye eklemişsiniz!\n\nListede var olan bir kelimeyi tekrar ekleyemezsiniz.')
                    else:
                        if len(tritem) == 0 or tritem.isspace() == True:
                            QtWidgets.QMessageBox.warning(
                                self, 'Uyarı', 'Lütfen kelimenin Türkçesini girin!')
                        else:
                            if len(item) == 0 or item.isspace() == True:
                                QtWidgets.QMessageBox.warning(
                                    self, 'Uyarı', 'Lütfen kelimenin İngilizcesini girin!')
                            else:
                                if len(tranlam) == 0 or tranlam.isspace() == True:
                                    QtWidgets.QMessageBox.warning(
                                        self, 'Uyarı', 'Lütfen kelimenin Türkçe ANLAMINI girin!')
                                else:

                                    if len(inganlam) == 0 or inganlam.isspace() == True:
                                        QtWidgets.QMessageBox.warning(
                                            self, 'Uyarı', 'Lütfen kelimenin İngilizce ANLAMINI girin!')

                                    else:
                                        tritems.append(tritem)
                                        items.append(item)
                                        items = list(dict.fromkeys(items))
                                        tritems = list(dict.fromkeys(tritems))
                                        tringitem = (item+" ["+tritem+"]")
                                        self.ui.listWidget_3.addItem(tringitem)
                                        self.ui.lineEdit_2.clear()
                                        self.ui.lineEdit.clear()
                                        turkceanlam = self.ui.plainTextEdit.toPlainText()
                                        ingilizceanlam = self.ui.plainTextEdit_2.toPlainText()

                                        trenanlam = tringitem+"\n\n"+"tr:"+"\n"+turkceanlam.capitalize()+"\n\n"+"en:"+"\n"+ingilizceanlam.capitalize()
                                        Window.tringanlam.append(trenanlam)
                                        self.ui.plainTextEdit.clear()
                                        self.ui.plainTextEdit_2.clear()

        self.kelimesayisi()
    def emailbilgi(self):

        try:
            for mail in excelemail:
                mail = mail
            for passw in excelpass:
                passw = passw
            try:
                authcookie = Office365('https://pirireisedutr.sharepoint.com/', username=mail,
                                       password=passw).GetCookies()
                i = 1
            except:
                i = 0
            if i == 1:
                site = Site('https://pirireisedutr.sharepoint.com/sites/GIN2006Listeler', version=Version.v2016,
                            authcookie=authcookie)
                now = datetime.now()
                dt_string = now.strftime("%d.%m.%Y %H.%M")
                filename = mail + " " + dt_string + ".txt"
                tringanlam = Window.tringanlam
                with open(filename, 'w') as file:
                    for line in tringanlam:
                        file.write(line)
                        file.write('\n''-----------------''\n')
                folder = site.Folder('Shared Documents/list')
                with open(filename, "rb") as file_obj:
                    file_as_string = file_obj.read()


                folder.upload_file(file_as_string, filename)
                try:
                    if os.path.exists(filename):
                        os.remove(filename)
                    else:
                        print("txt dosyası bulunamadı")
                except:
                    pass
            else:
                QtWidgets.QMessageBox.warning(
                    self, 'Hata', 'Beklenmeyen bir hata meydana geldi!\n\nLütfen ağ bağlantınızı kontrol edin.')
        except:
            pass

    def clearpanel(self):
        self.ui.lineEdit.clear()
        self.ui.plainTextEdit.clear()
        self.ui.plainTextEdit_2.clear()
        self.ui.lineEdit_2.clear()

    def findexcelkelime(self):
        items = Window.items
        Kelimeler = Window.Kelimelers
        item = self.ui.lineEdit.text()
        item = item.capitalize()
        if item in Kelimeler:
            QtWidgets.QMessageBox.warning(
                self, 'Uyarı', 'İstediğiniz kelime zaten Excel Dosyasına kayıtlı!\n\nExcel Dosyasında var olan bir kelimeyi tekrar ekleyemezsiniz.')
        else:
            self.addkelimelist()

    def findkelime(self):
        search_string = self.ui.lineEdit.text()
        match_items = self.ui.listWidget.findItems(search_string, Qt.MatchContains)
        for i in range(self.ui.listWidget.count()):
            it = self.ui.listWidget.item(i)
            it.setHidden(it not in match_items)
    birlesikkelime = []
    def listedenkaldir(self):
        birlesikkelime = Window.birlesikkelime
        if self.ui.listWidget_3.selectedItems():
            listItems = self.ui.listWidget_3.selectedItems()
            qm = QtWidgets.QMessageBox
            for item in listItems:
                ret = qm.question(self, 'Uyarı', "'"+item.text()+"'"+" kelimesini listeden kaldırmak istediğinizden emin misiniz?", qm.Yes | qm.No)
                if ret == qm.Yes:
                    if not listItems: return
                    for item in listItems:
                        self.ui.listWidget_3.takeItem(self.ui.listWidget_3.row(item))
                        index = self.ui.listWidget_3.currentRow()
                        item = item.text()
                        birlesikkelime.append(item)
                        parantezici = item[item.find("[") + 1:item.find("]")]
                        parantezdisi = item.split(' [')[0]
                        Window.items.remove(parantezdisi)
                        Window.tritems.remove(parantezici)


            self.check()
        else:
            pass
    site = []
    def yukle(self):
        if self.ui.listWidget_3.count() == 0:
            QtWidgets.QMessageBox.warning(
                self, 'Uyarı',
                "Listeye kelime eklemediniz, liste boşken Excel'e yükleme yapamazsınız!\n\nLütfen listeye kelime ekleyin.")
        else:
            qm = QtWidgets.QMessageBox

            ret = qm.question(self, 'Uyarı',
                              "Excel'e yüklenen kelimeler sonradan silinemez veya düzenlenemez!\n\nYükleme işlemini yapmadan önce eklemek istediğiniz kelimeleri kontrol etmeyi unutmayın.\n\nKelimeleri Excel'e kalıcı olarak yüklemek istediğinizden emin misiniz?",
                              qm.Yes | qm.No)
            if ret == qm.Yes:
                try:
                    for mail in excelemail:
                        mail = mail
                    for passw in excelpass:
                        passw = passw
                    try:
                        authcookie = Office365('https://pirireisedutr.sharepoint.com/', username=mail,
                                               password=passw).GetCookies()

                        i = 1
                    except:
                        i = 0
                    if i == 1:
                        site = Site('https://pirireisedutr.sharepoint.com/sites/GeminaatSzlkProjesiGrubu', version=Version.v2016,
                                    authcookie=authcookie)
                        Window.site.append(site)
                    folder = site.Folder('Shared Documents/Beta Testing')
                    self.download()
                    self.setlist()

                    wb = openpyxl.load_workbook("dict_data_setting.xlsx")
                    ws = wb['1']
                    tringanlam = Window.tringanlam
                    itemsTextList = [str(self.ui.listWidget_3.item(i).text()) for i in range(self.ui.listWidget_3.count())]

                    for kelime, anlam in zip(itemsTextList, tringanlam):
                        newRowLocation = ws.max_row + 1
                        ws.cell(column=1, row=newRowLocation, value=kelime)
                        ws.cell(column=2, row=newRowLocation, value=anlam)


                    wb.save(filename="dict_data_setting.xlsx")
                    wb.close()

                    with open("dict_data_setting.xlsx", "rb") as file_obj:
                        file_as_string = file_obj.read()
                    folder.upload_file(file_as_string, 'dict_data.xlsx')
                    time.sleep(2)
                    try:
                        for mail in excelemail:
                            mail = mail
                        for passw in excelpass:
                            passw = passw
                        try:
                            authcookie = Office365('https://pirireisedutr.sharepoint.com/', username=mail,
                                                   password=passw).GetCookies()
                            i = 1
                        except:
                            i = 0
                        if i == 1:
                            site = Site('https://pirireisedutr.sharepoint.com/sites/GIN2006Listeler',
                                        version=Version.v2016,
                                        authcookie=authcookie)
                            now = datetime.now()
                            dt_string = now.strftime("%d.%m.%Y %H.%M")
                            filename = mail + " " + dt_string + ".txt"
                            tringanlam = Window.tringanlam
                            with open(filename, 'w') as file:
                                for line in tringanlam:
                                    file.write(line)
                                    file.write('\n''-----------------''\n')
                            folder = site.Folder('Shared Documents/list')
                            with open(filename, "rb") as file_obj:
                                file_as_string = file_obj.read()

                            folder.upload_file(file_as_string, filename)
                        else:
                            QtWidgets.QMessageBox.warning(
                                self, 'Hata',
                                'Beklenmeyen bir hata meydana geldi!\n\nLütfen ağ bağlantınızı kontrol edin.')
                    except:
                        pass
                    time.sleep(2)
                    Window.items.clear()
                    self.ui.listWidget_3.clear()
                    self.kelimesayisi()
                    self.ui.listWidget.clear()
                    self.ui.listWidget.scrollToBottom()
                    self.updatekelimelist()
                    itemsTextList.clear()
                    tringanlam.clear()
                    self.msg = self.msg_wait()
                    self.msg.setStandardButtons(QMessageBox.Ok)
                    try:
                        if os.path.exists(filename):
                            os.remove(filename)
                        else:
                            print("txt dosyası bulunamadı")
                    except:
                        pass

                except:
                    pass

    def msg_wait(self):
        msg = QMessageBox()
        msg.setWindowIcon(QIcon("tick.png"))
        msg.setIconPixmap(QPixmap('tick.png'))
        msg.setText("\nYükleme işlemi başarıyla gerçekleştirildi.")
        msg.setWindowTitle("Başarılı")
        msg.setModal(False)
        msg.show()
        return msg
    def download(self):

        site = Window.site
        for i in site:
            i = i
        folder = i.Folder('Shared Documents/Beta Testing')
        file = folder.get_file('dict_data.xlsx')
        with open("dict_data.xlsx", "wb") as fh:
            fh.write(file)
        with open("dict_data.xlsx", "rb") as file_obj:
            file_as_string = file_obj.read()

    def check(self):
        self.kelimesayisi()
        try:
            sentences = Window.tringanlam
            words = Window.birlesikkelime

            indices = [[i for i, sentence in enumerate(sentences) if re.search('.+'.join(word.split()), sentence)] for word
                       in words]
            string_indices = str(indices)
            s = string_indices.replace('[', '')
            s = s.replace(']', '')
            string_indices = int(s)
            del sentences[string_indices]


            words.clear()
        except:
            pass
    def kelimesayisi(self):
        sayi = self.ui.listWidget_3.count()
        label = self.ui.label_6
        label.setText("Kelime Sayısı : "+str(sayi))
    def listedenkaldirduzenleme(self):
        birlesikkelime = Window.birlesikkelime
        if self.ui.listWidget_3.selectedItems():
            listItems = self.ui.listWidget_3.selectedItems()
            qm = QtWidgets.QMessageBox
            for item in listItems:
                if not listItems: return
                for item in listItems:
                    self.ui.listWidget_3.takeItem(self.ui.listWidget_3.row(item))
                    index = self.ui.listWidget_3.currentRow()

                    item = item.text()
                    birlesikkelime.append(item)
                    parantezici = item[item.find("[") + 1:item.find("]")]
                    # parantezdisi = item.split(' [')[0]
                    # Window.items.remove(parantezdisi)
                    Window.tritems.remove(parantezici)

            self.checkduzenlemesil()
    def checkduzenlemesil(self):
        try:
            sentences = Window.tringanlam
            words = Window.birlesikkelime

            indices = [[i for i, sentence in enumerate(sentences) if re.search('.+'.join(word.split()), sentence)] for word
                       in words]
            string_indices = str(indices)
            s = string_indices.replace('[', '')
            s = s.replace(']', '')
            string_indices = int(s)
            cumleler = sentences[string_indices]
            cumleler = str(cumleler)
            start = "tr:\n"
            end = "en:\n"
            words.clear()
            del sentences[string_indices]

        except:
            pass
    def checkduzenleme(self):
        birlesikkelime = Window.birlesikkelime
        if self.ui.listWidget_3.selectedItems():
            listItems = self.ui.listWidget_3.selectedItems()
            for item in listItems:
                if not listItems: return
                for item in listItems:
                    index = self.ui.listWidget_3.currentRow()

                    item = item.text()
                    birlesikkelime.append(item)
                    parantezici = item[item.find("[") + 1:item.find("]")]
                    parantezdisi = item.split(' [')[0]
                    self.ui.lineEdit.setText(parantezdisi)
                    self.ui.lineEdit_2.setText(parantezici)
                    try:
                        sentences = Window.tringanlam
                        words = Window.birlesikkelime

                        indices = [[i for i, sentence in enumerate(sentences) if re.search('.+'.join(word.split()), sentence)] for word
                                   in words]
                        string_indices = str(indices)
                        s = string_indices.replace('[', '')
                        s = s.replace(']', '')
                        string_indices = int(s)
                        cumleler = sentences[string_indices]
                        cumleler = str(cumleler)
                        start = "tr:\n"
                        end = "en:\n"
                        ing = cumleler.split("en:\n", 1)[1]

                        tr = cumleler[cumleler.find(start) + len(start):cumleler.rfind(end)]
                        self.ui.plainTextEdit.setPlainText(tr)
                        self.ui.plainTextEdit_2.setPlainText(ing)
                        words.clear()


                    except:
                        pass


if __name__ == '__main__':

    import sys
    app = QtWidgets.QApplication(sys.argv)
    login = Login()


    if login.exec_() == QtWidgets.QDialog.Accepted:
        window = Window()
        window.show()
        sys.exit(app.exec_())